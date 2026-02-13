from openpyxl import load_workbook, Workbook

# 1. Load the original workbook
wb = load_workbook('2feb.xlsx', data_only=True)
search_string = "Barrier gate opened"

# 2. Create a new workbook for the filtered results
new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Filtered Data"

print(f"Searching for: '{search_string}'...")

# 3. Iterate through every sheet and row
for sheet in wb.worksheets:
    for row in sheet.iter_rows(values_only=True): # values_only=True returns just the data
        # Check if the search string exists in any cell of the current row
        if any(cell and str(cell).strip() == search_string for cell in row):
            # Append the entire row of values to the new sheet
            new_sheet.append(row)
            print(f"Found and added row from Sheet: {sheet.title}")

# 4. Save the new workbook
new_wb.save('filtered_results.xlsx')
print("Search complete. Filtered data saved to 'filtered_results.xlsx'.")
