import glob
from openpyxl import load_workbook, Workbook

# 1. Setup
target_directory = './xlsx_outputs/*.xlsx'  # Change to your folder path
search_string = "Barrier gate opened"

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Consolidated Filtered Data"

# 2. Loop through all .xlsx files in the directory
for file_path in glob.glob(target_directory):
    print(f"Processing: {file_path}...")
    
    # Load each workbook
    wb = load_workbook(file_path, data_only=True)
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            # Check if search string exists in any cell of the row
            if any(cell and str(cell).strip() == search_string for cell in row):
                new_sheet.append(row)

# 3. Save the single consolidated file
new_wb.save('all_filtered_results.xlsx')
print("\nSuccess! All matching rows saved to 'all_filtered_results.xlsx'.")
